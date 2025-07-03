import logging
import inspect
from openpyxl import load_workbook

class utils:
    @staticmethod
    def logger(loglevel=logging.DEBUG):
        """
        Creates and returns a logger instance with the given log level.

        - Dynamically sets the logger name based on the calling function.
        - Writes logs to a file with timestamp, logger name, and message.
        - Log file is overwritten on each execution (`mode='w'`).

        Args:
            loglevel (int): Logging level (e.g., logging.INFO, logging.ERROR)

        Returns:
            logger (logging.Logger): Configured logger object
        """
        # Gets the name of the calling method to use as the logger name
        log_name = inspect.stack()[1][3]
        logger = logging.getLogger(log_name)
        logger.setLevel(loglevel)

        # Format for log messages
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

        # FileHandler to write logs to specified log file
        fh = logging.FileHandler("D:\\Automation\\Rubick_Assignment\\Flow\\Logs\\Logs.log", mode='w')
        fh.setFormatter(formatter)

        # Avoid adding multiple handlers if already added
        if not logger.hasHandlers():
            logger.addHandler(fh)

        return logger

    @staticmethod
    def read_xlsx(file_name, sheet):
        """
        Reads an Excel sheet and returns data from all rows except the header row.

        Args:
            file_name (str): Full path to the Excel (.xlsx) file.
            sheet (str): Name of the sheet to read.

        Returns:
            datalist (list of lists): Row-wise data from Excel sheet.
        """
        datalist = []  # To store data rows
        try:
            # Load workbook and access specified sheet
            wb = load_workbook(filename=file_name)
            sh = wb[sheet]

            row_ct = sh.max_row
            col_ct = sh.max_column

            # Loop from second row to skip headers
            for i in range(2, row_ct + 1):
                row = []
                for j in range(1, col_ct + 1):
                    row.append(sh.cell(row=i, column=j).value)
                datalist.append(row)

            return datalist

        except Exception as ex:
            logging.error(f"Error while reading Excel file: {ex}")
            return []  # Return empty list on failure
